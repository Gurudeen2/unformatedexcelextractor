import openpyxl
from time import gmtime, strftime
import os
import shutil
import datetime
import json
from array import *
import re
import cx_Oracle

# funcrion for dob

def nam(p):
    dnameindive=''
    dnameindives=p.split(' ')
    if len(dnameindives) > 1:
        if len(dnameindives) == 2:
            return dnameindive
        if len(dnameindives) <= len(dnameindives):
            return dnameindives[1]
        return dnameindive
    return dnameindive


def dob(x): 
    resu = re.findall(r'\d{4}-\d{2}-\d{2}' , x)
    ic4prod='' 
    if len(resu)>=1:
        return resu[0]
    return ic4prod

#REMOVING NON ALPHABETS FROM THE FULL NAME
def without_non_alphabets(x):
    pattern = re.compile("\W")
    y = pattern.sub('',x)
    return y

def without_non_alphabet(x):
    pattern = re.compile("\W")
    y = pattern.sub(' ',x)
    return y

def remnone(x):
    return re.sub("None","", x)
# whatch id

def watchId(fullnames):

    strs=fullnames.split(' ')
    strss1=''
    for char in strs:
        if len(char)>1:
            if len(char)==5:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4]))
                return strss1
            if len(char)==6:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5]))
                return strss1
            if len(char)==7:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+ord(char[6]))
                return strss1
            if len(char)==8:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7]))
                return strss1
            if len(char)==9:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8]))
                return strss1
            if len(char)==10:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9]))
                return strss1
            if len(char)==11:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9])+ord(char[10]))
                return strss1
            if len(char)==12:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9])+ord(char[10])+ord(char[11]))
                return strss1
            if len(char)==13:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9])+ord(char[10])+ord(char[11])+ord(char[12]))
                return strss1
            if len(char)==14:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9])+ord(char[10])+ord(char[11])+ord(char[12])+ord(char[13]))
                return strss1
            if len(char)>=14:
                strss1=str(ord(char[0])+ord(char[1])+ord(char[2])+ord(char[3])+ord(char[4])+ord(char[5])+
                            ord(char[6])+ord(char[7])+ord(char[8])+ord(char[9])+ord(char[10])+ord(char[11])+ord(char[12])+ord(char[13]))
                return strss1
                        
    return strss1

#CONDITIONAL STATEMENT REMOVE ANY NULL FROM OUR DATA AND MAKE IT EMPTY STRINGS
def removing_none(x):
    if(x=='None'):
        x=''
    return x
def without_numeral(x):
    pattern = re.compile("[\d]+")
    y = pattern.sub('',x)
    return y


def last_name(x):
    dnameindive=''
    dnameindives=x.split(' ')
#    ['akk', '', 'fgfg']
    if len(dnameindives) > 1:
        if len(dnameindives) == 2:
            return dnameindives[1]
        if len(dnameindives) <= len(dnameindives):
            return dnameindives[2]
        return dnameindive
    return dnameindive


def middle_name(x):
    dnameindive=''
    dnameindives=x.split(' ')
    if len(dnameindives) > 1:
        if len(dnameindives) == 2:
            return dnameindive
        if len(dnameindives) <= len(dnameindives):
            return dnameindives[1]
        return dnameindive
    return dnameindive


def correctdate(dates):
    return dates.replace(" 00:00:00", "")

def removedoublehyphen(x):
    return x.replace("--", "-") 

def removeapro(x):
    return x.replace("'", "")

def replaceslash(x):
    return x.replace("/", "-")

def replacespace(x):
    return x.replace(" ", "")


# impor the xlsx file into the code
try:
   source = "C:\\Users\\HP\\Documents\\ReW\\PEP.xlsx"
   destination = "C:\\Users\\HP\\Documents\\ReW\\new"
   shutil.move(source, destination)
except Exception as e:
        print('Failed to move %s. Reason: %s' % (e))
# Give the Directory of the file 
my_directory = "C:\\Users\\HP\\Documents\\ReW\\new"

d =datetime.date.today()
t =strftime("%H:%M:%S", gmtime())

con = cx_Oracle.connect('U_IC4INDEP/c4@localhost:1521/orcl')
cursor = con.cursor()
with os.scandir(my_directory) as entries:
    for entry in entries:
        ext_name=entry.name
        if ext_name.endswith('.xlsx'):
    
         wb_obj = openpyxl.load_workbook(entry) 
  
         sheet_obj = wb_obj.active 
         m_row = sheet_obj.max_row 
         m_col = sheet_obj.max_column
         
         result = []
         transactions = {}
         summary = {}

         v_counter = 0
        #  print("XXXX No of rows: ", m_row)
        #  print("YYYY No of Col: ",m_col)

         
         
         relevantRows = []
        
         for i in range(2, m_row + 1):
            row = []
            if sheet_obj.cell(row=i, column=1).value is None:
                continue

            data1 = sheet_obj.cell(row=i, column=1).value
         
            part1 = data1

            # print("part1: ", part1)
            for k in range(1, m_col+1):  
               cell_obj = sheet_obj.cell(row = i, column = k) 
               row.append(cell_obj.value)
               continue
            with open('tttt.txt','a') as f:
                f.write(str(row))
                f.write('\n')

            todaysdatetime = datetime.datetime.now()
            todaysdate = datetime.date.today()
            format = "%d/%m/%Y"
            radproAltDOB=''	
            radproBlacklistedId	=''
            radproReportedBy=''	
            radproPortName=''	
            radproPortType =''	
            radproWhiteFlag	=''
            radproFlagDate	=''
            radproFlagTime	=''
            radproReason = ''	
            radproFlagBy = ''	
            radproRemarks =''	
            radproTags =''	
            radproProfession=str(row[12])
            watchIDs=''
            radproAssociateId=''	
            radproWatchlistId =''
            radproTitle	=str(row[4])
            radproFirstName	= str(row[2])
            radproMiddleName =str(row[3])
            radproSurname	= str(row[1])
            radproMaidenName = str(row[19])	
            radproOtherNames = ''
            radproFullName =str(row[2])+' '+str(row[3])+' '+str(row[1])	
            radproRelName = ''	
            radproType	=''
            radproAction=''	
            radproActionDate=''	
            radproGender = str(row[6])	
            radproStatus=''	
            radproDOB =	correctdate (str(row[7]))
            radproPOB = str(row[24])
            radproAddress=str(row[11])
            radproCity	=''
            radproCounty=''	
            radproPostalCode=''	
            radproState	=''
            radproCountry	='NIGERIA'
            radproCountryCode='NG'	
            radproContinent	= 'AFRICA'
            radproCategory	= 'PEP'
            radproSubCategory =''	
            radproWatchType	='INDIVIDUAL'
            radproBranch = ''
            radproCluster	=''
            radproZone	=''
            radproRegion =''	
            radproImageLink	=''
            radproUrl	=''
            radproTier	=''
            radproRisk =''
            radproTierRank=''	
            radproDescription=''	
            radproLanguage='en'	
            radproRecordDate = d	
            radproRecordTime = t	
            radproOperator='SYSTEM'	
            radproOperation	='CREATE'
            radproRecordCounter = 1
            array1 = ['Dr.(Mrs.)','Dr','Dr.','Mrs.','Oba','late','Chief','Late',
                     'Justice','Dr Mrs',' Mr',' Ms.',' Mr.',' Ms','Chief (Mrs)','Miss','Miss','Engr','Engr.',
                     'Prince','Hon.','Alhaji','Princess','Professor','Prophetess','Alhaja','Mrs','(Mrs.)','Hon.(Mrs.)']
           
            radproWorkstation= "AUTOLOAD"
            titlearray =['Hon.', 'Chief', 'Alhaja', 'is', 'the', 'late', 'Mrs.', 
            'Rev.', 'Prof.', 'Princess', 'Gov.', 'Mr.', 'Dr', 'Olori', 'Late', 'Miss', '(Ex-Wife)', 'Oloori',
             'Barr.', 'Comm', 'Mrs', '(Mrs.)']
            sib_title=''
            sib_firstName=''
            sib_middleName=''
            sib_lastName=''
            Note=str(row[29])
            business_interest=str(row[21])
            email_address=str(row[28])
            Description=str(row[10])
            profession=str(row[12])
            relNames=''
            firstNames=''
            spouse=str(row[15])
            children=str(row[16])
            sibling=str(row[17])
            parent=str(row[18])
            listCode='PEP'
            startDate=str(row[26])
            endDate=''
            name=''
            status=''
            description1=str(row[27])
            description2=str(row[25])
            title=''
            radproAssociateId=''
            radproAssociateIds=''
            radproAssociateIdss=''
            radproAssociateIdsss=''
            radproAssociateIdssss=''
            radproAs=''
            radproAsso=''
            cib_title=''
            cib_firstName=''
            cib_middleName=''
            cib_lastName=''
            scib_title=''
            scib_firstName=''
            scib_middleName=''
            scib_lastName=''
            radproPOB=without_numeral(radproPOB)
            ic4proFullName=without_non_alphabets(radproFullName)
            watchID=watchId(ic4proFullName)
            radproWatchlistId = radproCategory+"-"+radproFirstName.upper() +'-'+ watchID
            
            
            PersonKeys = (
                radproAssociateId
                    
                    ) 
            # siblings
            relNames='siblings'
            if(sibling=='None'):
               sibling=''
            else:
                find_it = "and"
                repl_it = ","
                sibling = sibling.replace(find_it, repl_it)
                sibling=sibling.split(",")
                if len(sibling) > 1:
                    for i in sibling:
                        j=i
                        i=i.split(' ')
                        if len(i) > 1:
                            sib_firstName=i[0]
                            sib_middleName=i[1]
                            if sib_firstName in array1:
                                sib_title=i[0]
                                sib_firstName=i[1]
                            else:
                                sib_firstName=i[0]

                            if sib_firstName=='':
                                sib_firstName=i[1]
                            else:
                                sib_firstName= sib_firstName
                        else:
                            sib_firstName=i
                    
                        sib_middleName=str(sib_middleName).split(' ')
                        if len(sib_middleName)>1:
                            sib_middleName=sib_middleName[0]
                            sib_lastName=sib_middleName[1]
                        else:
                            sib_lastName=sib_middleName[0]

                        sib_middleName=str(sib_middleName).removeprefix("['")
                        sib_middleName=str(sib_middleName).removesuffix("']")
                        sib_middleName=str(sib_middleName).removeprefix('["')
                        sibling=str(sib_middleName).removesuffix('"]')  

                        sib_lastName=str(sib_lastName).removeprefix("['")
                        sib_lastName=str(sib_lastName).removesuffix("']")
                        sib_lastName=str(sib_lastName).removeprefix('["')
                        sib_lastName=str(sib_lastName).removesuffix('"]')

                        sib_firstName=str(sib_firstName).removeprefix("['")
                        sib_firstName=str(sib_firstName).removesuffix("']")
                        sib_firstName=str(sib_firstName).removeprefix('["')
                        sib_firstName=str(sib_firstName).removesuffix('"]')
                        
                        j=str(j).removeprefix("['")
                        j=str(j).removesuffix("']")
                        j=str(j).removeprefix('["')
                        j=str(j).removesuffix('"]')
                    
                    
                        val = removing_none( without_non_alphabets (j))
                        watchIDs=watchId(val)
                        middlename = middle_name(j)
                        lastname = last_name(j)
                        radproAssociateId =radproCategory+"-"+ lastname+"-"+radproFirstName +'-'+ watchID
                        if str(sib_middleName)==str(sib_lastName):
                            sib_middleName=''

                        assosibling=(
                        replacespace (removedoublehyphen (radproAssociateId.upper())),
                        replacespace(removedoublehyphen(radproWatchlistId.upper())),	
                        removing_none(radproTitle),
                        removing_none(sib_firstName),
                        removing_none(middlename),
                        removing_none(radproSurname),
                        removing_none(radproMaidenName),	
                        removing_none(radproOtherNames),
                        str(j),	
                        removing_none(relNames),	
                        removing_none(radproType),	
                        removing_none(radproAction),	
                        removing_none(radproActionDate),	
                        removing_none(radproGender),	
                        removing_none(radproStatus),	
                        removing_none(radproDOB),	
                        removing_none(radproPOB),	
                        removing_none(radproAddress),	
                        removing_none(radproCity),	
                        removing_none(radproCounty),	
                        removing_none(radproPostalCode),	
                        removing_none(radproState),	
                        removing_none(radproCountry),	
                        removing_none(radproCountryCode),	
                        removing_none(radproContinent),	
                        removing_none(radproCategory),	
                        removing_none(radproSubCategory),	
                        removing_none(radproWatchType),	
                        removing_none(radproBranch),	
                        removing_none(radproCluster),	
                        removing_none(radproZone),
                        removing_none(radproRegion),	
                        removing_none(radproImageLink),	
                        removing_none(radproUrl),	
                        removing_none(radproTier),	
                        removing_none(radproRisk),	
                        removing_none(radproTierRank),	
                        removing_none(radproDescription),	
                        removing_none(radproLanguage),	
                        str(radproRecordDate),	
                        removing_none(radproRecordTime),	
                        removing_none(radproOperator),	
                        removing_none(radproOperation),	
                        removing_none(radproWorkstation),	
                        removing_none(radproRecordCounter))
                        with open('textt.txt', 'a') as f:
                            f.write(str(str(replacespace (removedoublehyphen (radproAssociateId.upper())))))
                            f.write('\n')
                        cursor.execute('''
                        SELECT * FROM U_IC4INDEP.PEP_ASSOCIATES WHERE RADPROASSOCIATEID='{0}'
                    '''.format(str(removeapro (replacespace (removedoublehyphen (radproAssociateId.upper()))))))
                        row = cursor.fetchone()
                        if not row:
                            cursor.execute('''
                            INSERT INTO "U_IC4INDEP"."PEP_ASSOCIATES" 
                            (
                                    "RADPROASSOCIATEID" , 
                                    "RADPROWATCHLISTID" , 
                                    "RADPROTITLE", 
                                    "RADPROFIRSTNAME" , 
                                    "RADPROMIDDLENAME" , 
                                    "RADPROSURNAME" , 
                                    "RADPROMAIDENNAME" , 
                                    "RADPROOTHERNAMES" , 
                                    "RADPROFULLNAME" , 
                                    "RADPRORELNAME" , 
                                    "RADPROTYPE" , 
                                    "RADPROACTION" , 
                                    "RADPROACTIONDATE" , 
                                    "RADPROGENDER" , 
                                    "RADPROSTATUS" , 
                                    "RADPRODOB" , 
                                    "RADPROPOB", 
                                    "RADPROADDRESS", 
                                    "RADPROCITY" , 
                                    "RADPROCOUNTY" , 
                                    "RADPROPOSTALCODE" , 
                                    "RADPROSTATE" , 
                                    "RADPROCOUNTRY" , 
                                    "RADPROCOUNTRYCODE" , 
                                    "RADPROCONTINENT" , 
                                    "RADPROCATEGORY" , 
                                    "RADPROSUBCATEGORY" , 
                                    "RADPROWATCHTYPE" , 
                                    "RADPROBRANCH" , 
                                    "RADPROCLUSTER" , 
                                    "RADPROZONE" , 
                                    "RADPROREGION" , 
                                    "RADPROIMAGELINK" , 
                                    "RADPROURL" , 
                                    "RADPROTIER" , 
                                    "RADPRORISK" , 
                                    "RADPROTIERRANK" , 
                                    "RADPRODESCRIPTION" , 
                                    "RADPROLANGUAGE" , 
                                    "RADPRORECORDDATE" , 
                                    "RADPRORECORDTIME" , 
                                    "RADPROOPERATOR" , 
                                    "RADPROOPERATION" , 
                                    "RADPROWORKSTATION" , 
                                    "RADPRORECORDCOUNTER"

                            )
                            VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}',
                                    '{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
                                    '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}','{30}','{31}','{32}',
                                    '{33}','{34}','{35}','{36}','{37}','{38}','{39}','{40}','{41}','{42}','{43}',
                                    '{44}'
                            )
                            '''.format(
                            removeapro(str(assosibling[0])), removeapro(str(assosibling[1])), removeapro(str(assosibling[2])),
                            removeapro(str(assosibling[3])), removeapro(str(assosibling[4])), removeapro(str(assosibling[5])), 
                            removeapro(str(assosibling[6])), removeapro(str(assosibling[7])), removeapro(str(assosibling[8])), 
                            removeapro(str(assosibling[9])), removeapro(str(assosibling[10])), removeapro(str(assosibling[11])), 
                            removeapro(str(assosibling[12])), removeapro(str(assosibling[13])), removeapro(str(assosibling[14])), 
                            replaceslash(removeapro(str(assosibling[15]))), removeapro(str(assosibling[16])),
                            removeapro(str(assosibling[17])),removeapro(str(assosibling[18])),removeapro(str(assosibling[19])),
                            removeapro(str(assosibling[20])),removeapro(str(assosibling[21])),removeapro(str(assosibling[22])),removeapro(str(assosibling[23])),removeapro(str(assosibling[24])),
                            removeapro(str(assosibling[25])),
                            removeapro(str(assosibling[26])),removeapro(str(assosibling[27])),removeapro(str(assosibling[28])),
                            removeapro(str(assosibling[29])),removeapro(str(assosibling[30])),removeapro(str(assosibling[31])),
                            removeapro(str(assosibling[32])),removeapro(str(assosibling[33])),removeapro(str(assosibling[34])),
                            removeapro(str(assosibling[35])),removeapro(str(assosibling[36])),removeapro(str(assosibling[37])),
                            removeapro(str(assosibling[38])),removeapro(str(assosibling[39])),removeapro(str(assosibling[40])),
                            removeapro(str(assosibling[41])),removeapro(str(assosibling[42])),removeapro(str(assosibling[43])),
                            removeapro(str(assosibling[44]))                        
                            
                            ))
                            con.commit()
                else:
                    sibling=str(sibling).removeprefix("['")
                    sibling=sibling.removesuffix("']")
                    sibling=str(sibling).removeprefix('["')
                    sibling=sibling.removesuffix('"]')  
                    sibb=sibling
                    sibling=sibling.split(' ')
                    if len(sibling) > 1:
                        sib_firstName=sibling[0]
                        sib_middleName=sibling[1]
                        if sib_firstName in array1:
                            sib_title=sibling[0]
                            sib_firstName=sibling[1]
                        else:
                            sib_firstName=sibling[0]
                        sib_middleName=str(sib_middleName).split(' ')
                        if len(sib_middleName)>1:
                            sib_middleName=sib_middleName[0]
                            sib_lastName=sib_middleName[1]
                        else:
                            sib_lastName=sib_middleName[0]
                        sib_middleName=str(sib_middleName).removeprefix("['")
                        sib_middleName=str(sib_middleName).removesuffix("']")
                        sib_middleName=str(sib_middleName).removeprefix('["')
                        sibling=str(sib_middleName).removesuffix('"]')  

                        sib_lastName=str(sib_lastName).removeprefix("['")
                        sib_lastName=str(sib_lastName).removesuffix("']")
                        sib_lastName=str(sib_lastName).removeprefix('["')
                        sib_lastName=str(sib_lastName).removesuffix('"]')

                        sib_firstName=str(sib_firstName).removeprefix("['")
                        sib_firstName=str(sib_firstName).removesuffix("']")
                        sib_firstName=str(sib_firstName).removeprefix('["')
                        sib_firstName=str(sib_firstName).removesuffix('"]')

                        sibb=str(sibb).removeprefix("['")
                        sibb=str(sibb).removesuffix("']")
                        sibb=str(sibb).removeprefix('["')
                        sibb=str(sibb).removesuffix('"]')

                        middlename2 = middle_name(sibb)
                        lastname2 = last_name(sibb)
                        watchIDss=watchId(sibb)

                        radproAssociateId =radproCategory+"-"+ middlename2+"-"+radproFirstName +'-'+ watchID                        
                        radproWatchlistId =radproCategory+"-"+radproFirstName +'-'+ watchID                        
                        
                        if str(sib_middleName)==str(sib_lastName):
                            sib_middleName=''
    
                        assosibling2=(
                        replacespace(removedoublehyphen (radproAssociateId.upper())),
                        replacespace(removedoublehyphen(radproWatchlistId.upper())),	
                        removing_none(sib_title),
                        removing_none(sib_firstName),
                        removing_none(middlename2),
                        removing_none(lastname2),
                        removing_none(radproMaidenName),	
                        removing_none(radproOtherNames),
                        str(sibb),	
                        removing_none(relNames),	
                        removing_none(radproType),	
                        removing_none(radproAction),	
                        removing_none(radproActionDate),	
                        removing_none(radproGender),	
                        removing_none(radproStatus),	
                        removing_none(radproDOB),	
                        removing_none(radproPOB),	
                        removing_none(radproAddress),	
                        removing_none(radproCity),	
                        removing_none(radproCounty),	
                        removing_none(radproPostalCode),	
                        removing_none(radproState),	
                        removing_none(radproCountry),	
                        removing_none(radproCountryCode),	
                        removing_none(radproContinent),	
                        removing_none(radproCategory),	
                        removing_none(radproSubCategory),	
                        removing_none(radproWatchType),	
                        removing_none(radproBranch),	
                        removing_none(radproCluster),	
                        removing_none(radproZone),
                        removing_none(radproRegion),	
                        removing_none(radproImageLink),	
                        removing_none(radproUrl),	
                        removing_none(radproTier),	
                        removing_none(radproRisk),	
                        removing_none(radproTierRank),	
                        removing_none(radproDescription),	
                        removing_none(radproLanguage),	
                        str(radproRecordDate),	
                        removing_none(radproRecordTime),	
                        removing_none(radproOperator),	
                        removing_none(radproOperation),	
                        removing_none(radproWorkstation),	
                        removing_none(radproRecordCounter)
                        )
                        cursor.execute('''
                        SELECT * FROM U_IC4INDEP.PEP_ASSOCIATES WHERE RADPROASSOCIATEID='{0}'
                    '''.format(str(replacespace(removedoublehyphen(radproAssociateId.upper())))))
                        row = cursor.fetchone()
                        if not row:
                            cursor.execute('''
                        INSERT INTO "U_IC4INDEP"."PEP_ASSOCIATES" 
                        (
                                "RADPROASSOCIATEID" , 
                                "RADPROWATCHLISTID" , 
                                "RADPROTITLE", 
                                "RADPROFIRSTNAME" , 
                                "RADPROMIDDLENAME" , 
                                "RADPROSURNAME" , 
                                "RADPROMAIDENNAME" , 
                                "RADPROOTHERNAMES" , 
                                "RADPROFULLNAME" , 
                                "RADPRORELNAME" , 
                                "RADPROTYPE" , 
                                "RADPROACTION" , 
                                "RADPROACTIONDATE" , 
                                "RADPROGENDER" , 
                                "RADPROSTATUS" , 
                                "RADPRODOB" , 
                                "RADPROPOB", 
                                "RADPROADDRESS", 
                                "RADPROCITY" , 
                                "RADPROCOUNTY" , 
                                "RADPROPOSTALCODE" , 
                                "RADPROSTATE" , 
                                "RADPROCOUNTRY" , 
                                "RADPROCOUNTRYCODE" , 
                                "RADPROCONTINENT" , 
                                "RADPROCATEGORY" , 
                                "RADPROSUBCATEGORY" , 
                                "RADPROWATCHTYPE" , 
                                "RADPROBRANCH" , 
                                "RADPROCLUSTER" , 
                                "RADPROZONE" , 
                                "RADPROREGION" , 
                                "RADPROIMAGELINK" , 
                                "RADPROURL" , 
                                "RADPROTIER" , 
                                "RADPRORISK" , 
                                "RADPROTIERRANK" , 
                                "RADPRODESCRIPTION" , 
                                "RADPROLANGUAGE" , 
                                "RADPRORECORDDATE" , 
                                "RADPRORECORDTIME" , 
                                "RADPROOPERATOR" , 
                                "RADPROOPERATION" , 
                                "RADPROWORKSTATION" , 
                                "RADPRORECORDCOUNTER"

                        )
                        VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}',
                                '{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
                                '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}','{30}','{31}','{32}',
                                '{33}','{34}','{35}','{36}','{37}','{38}','{39}','{40}','{41}','{42}','{43}',
                                '{44}'
                        )
                        '''.format(
                        removeapro(str(assosibling2[0])), removeapro(str(assosibling2[1])), removeapro(str(assosibling2[2])),
                        removeapro(str(assosibling2[3])), removeapro(str(assosibling2[4])), removeapro(str(assosibling2[5])), 
                        removeapro(str(assosibling2[6])), removeapro(str(assosibling2[7])), removeapro(str(assosibling2[8])), 
                        removeapro(str(assosibling2[9])), removeapro(str(assosibling2[10])), removeapro(str(assosibling2[11])), 
                        removeapro(str(assosibling2[12])), removeapro(str(assosibling2[13])), removeapro(str(assosibling2[14])), 
                        replaceslash(removeapro(str(assosibling2[15]))), removeapro(str(assosibling2[16])),
                        removeapro(str(assosibling2[17])),removeapro(str(assosibling2[18])),removeapro(str(assosibling2[19])),
                        removeapro(str(assosibling2[20])),removeapro(str(assosibling2[21])),removeapro(str(assosibling2[22])),removeapro(str(assosibling2[23])),removeapro(str(assosibling2[24])),
                        removeapro(str(assosibling2[25])),
                        removeapro(str(assosibling2[26])),removeapro(str(assosibling2[27])),removeapro(str(assosibling2[28])),
                        removeapro(str(assosibling2[29])),removeapro(str(assosibling2[30])),removeapro(str(assosibling2[31])),
                        removeapro(str(assosibling2[32])),removeapro(str(assosibling2[33])),removeapro(str(assosibling2[34])),
                        removeapro(str(assosibling2[35])),removeapro(str(assosibling2[36])),removeapro(str(assosibling2[37])),
                        removeapro(str(assosibling2[38])),removeapro(str(assosibling2[39])),removeapro(str(assosibling2[40])),
                        removeapro(str(assosibling2[41])),removeapro(str(assosibling2[42])),removeapro(str(assosibling2[43])),
                        removeapro(str(assosibling2[44]))                        
                        
                        ))
                            con.commit()

            relNames='children'
            if(children=='None'):
                children=''
            else:
                find_it = "and"
                repl_it = ","
                children = children.replace(find_it, repl_it)
                children=children.split(",")
                if len(children) > 1:
                    for i in children:
                        k=i
                        i=i.split(' ')
                    if len(i) > 1:
                        cib_firstName=i[0]
                        cib_middleName=i[1]
                        if cib_firstName in array1:
                            cib_title=i[0]
                            cib_firstName=i[1]
                        else:
                            cib_firstName=i[0]
                        if cib_firstName=='':
                            cib_firstName=i[1]
                        else:
                            cib_firstName=cib_firstName
                    else:
                        cib_firstName=i
                    cib_middleName=str(cib_middleName).split(' ')
                    if len(cib_middleName)>1:
                        cib_middleName=cib_middleName[0]
                        cib_lastName=cib_middleName[1]
                    else:
                        cib_lastName=cib_middleName[0]
                    
                    cib_middleName=str(cib_middleName).removeprefix("['")
                    cib_middleName=str(cib_middleName).removesuffix("']")
                    cib_middleName=str(cib_middleName).removeprefix('["')
                    children=str(cib_middleName).removesuffix('"]')  

                    cib_lastName=str(cib_lastName).removeprefix("['")
                    cib_lastName=str(cib_lastName).removesuffix("']")
                    cib_lastName=str(cib_lastName).removeprefix('["')
                    cib_lastName=str(cib_lastName).removesuffix('"]')

                    cib_firstName=str(cib_firstName).removeprefix("['")
                    cib_firstName=str(cib_firstName).removesuffix("']")
                    cib_firstName=str(cib_firstName).removeprefix('["')
                    cib_firstName=str(cib_firstName).removesuffix('"]')

                    k=str(k).removeprefix("['")
                    k=str(k).removesuffix("']")
                    k=str(k).removeprefix('["')
                    k=str(k).removesuffix('"]')

           

                    middlename3 =middle_name(k)
            #          #print(ic4proFirstNames)


                    lastname3 =last_name(k)
            #          #print(ic4proSurnames)

                    watchIDd=watchId(k)
                    radproAssociateId =radproCategory+"-"+ without_non_alphabets (str(cib_firstName))+"-"+radproFirstName +'-'+ watchID
                    radproWatchlistId =radproCategory+"-"+ "-"+radproFirstName +'-'+ watchID

                    if str(cib_middleName)==str(cib_lastName):
                        cib_middleName=''

                    assochildren=(
                        replacespace(removedoublehyphen (radproAssociateId.upper())),
                        replacespace(removedoublehyphen(radproWatchlistId.upper())),	
                        removing_none(cib_title),
                        removing_none(without_non_alphabets (str(cib_firstName))),
                        removing_none(middlename3),
                        removing_none(lastname3),
                        removing_none(radproMaidenName),	
                        removing_none(radproOtherNames),
                        without_non_alphabet(str(k)),	
                        relNames,	
                        removing_none(radproType),	
                        removing_none(radproAction),	
                        removing_none(radproActionDate),	
                        removing_none(radproGender),	
                        removing_none(radproStatus),	
                        replaceslash(removing_none(radproDOB)),	
                        removing_none(radproPOB),	
                        removing_none(radproAddress),	
                        removing_none(radproCity),	
                        removing_none(radproCounty),	
                        removing_none(radproPostalCode),	
                        removing_none(radproState),	
                        removing_none(radproCountry),	
                        removing_none(radproCountryCode),	
                        removing_none(radproContinent),	
                        removing_none(radproCategory),	
                        removing_none(radproSubCategory),	
                        removing_none(radproWatchType),	
                        removing_none(radproBranch),	
                        removing_none(radproCluster),	
                        removing_none(radproZone),
                        removing_none(radproRegion),	
                        removing_none(radproImageLink),	
                        removing_none(radproUrl),	
                        removing_none(radproTier),	
                        removing_none(radproRisk),	
                        removing_none(radproTierRank),	
                        removing_none(radproDescription),	
                        removing_none(radproLanguage),	
                        str(radproRecordDate),	
                        removing_none(radproRecordTime),	
                        removing_none(radproOperator),	
                        removing_none(radproOperation),	
                        removing_none(radproWorkstation),	
                        removing_none(radproRecordCounter)
                    )
                    cursor.execute('''
                        SELECT * FROM U_IC4INDEP.PEP_ASSOCIATES WHERE RADPROASSOCIATEID='{0}'
                    '''.format(str(replacespace(removedoublehyphen(radproAssociateId.upper())))))
                    row = cursor.fetchone()
                    if not row:
                        cursor.execute('''
                        INSERT INTO "U_IC4INDEP"."PEP_ASSOCIATES" 
                        (
                                "RADPROASSOCIATEID" , 
                                "RADPROWATCHLISTID" , 
                                "RADPROTITLE", 
                                "RADPROFIRSTNAME" , 
                                "RADPROMIDDLENAME" , 
                                "RADPROSURNAME" , 
                                "RADPROMAIDENNAME" , 
                                "RADPROOTHERNAMES" , 
                                "RADPROFULLNAME" , 
                                "RADPRORELNAME" , 
                                "RADPROTYPE" , 
                                "RADPROACTION" , 
                                "RADPROACTIONDATE" , 
                                "RADPROGENDER" , 
                                "RADPROSTATUS" , 
                                "RADPRODOB" , 
                                "RADPROPOB", 
                                "RADPROADDRESS", 
                                "RADPROCITY" , 
                                "RADPROCOUNTY" , 
                                "RADPROPOSTALCODE" , 
                                "RADPROSTATE" , 
                                "RADPROCOUNTRY" , 
                                "RADPROCOUNTRYCODE" , 
                                "RADPROCONTINENT" , 
                                "RADPROCATEGORY" , 
                                "RADPROSUBCATEGORY" , 
                                "RADPROWATCHTYPE" , 
                                "RADPROBRANCH" , 
                                "RADPROCLUSTER" , 
                                "RADPROZONE" , 
                                "RADPROREGION" , 
                                "RADPROIMAGELINK" , 
                                "RADPROURL" , 
                                "RADPROTIER" , 
                                "RADPRORISK" , 
                                "RADPROTIERRANK" , 
                                "RADPRODESCRIPTION" , 
                                "RADPROLANGUAGE" , 
                                "RADPRORECORDDATE" , 
                                "RADPRORECORDTIME" , 
                                "RADPROOPERATOR" , 
                                "RADPROOPERATION" , 
                                "RADPROWORKSTATION" , 
                                "RADPRORECORDCOUNTER"

                        )
                        VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}',
                                '{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
                                '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}','{30}','{31}','{32}',
                                '{33}','{34}','{35}','{36}','{37}','{38}','{39}','{40}','{41}','{42}','{43}',
                                '{44}'
                        )
                        '''.format(
                        removeapro(str(assochildren[0])), removeapro(str(assochildren[1])), removeapro(str(assochildren[2])),
                        removeapro(str(assochildren[3])), removeapro(str(assochildren[4])), removeapro(str(assochildren[5])), 
                        removeapro(str(assochildren[6])), removeapro(str(assochildren[7])), removeapro(str(assochildren[8])), 
                        removeapro(str(assochildren[9])), removeapro(str(assochildren[10])), removeapro(str(assochildren[11])), 
                        removeapro(str(assochildren[12])), removeapro(str(assochildren[13])), removeapro(str(assochildren[14])), 
                        replaceslash(removeapro(str(assochildren[15]))), removeapro(str(assochildren[16])),
                        removeapro(str(assochildren[17])),removeapro(str(assochildren[18])),removeapro(str(assochildren[19])),
                        removeapro(str(assochildren[20])),removeapro(str(assochildren[21])),removeapro(str(assochildren[22])),removeapro(str(assochildren[23])),removeapro(str(assochildren[24])),
                        removeapro(str(assochildren[25])),
                        removeapro(str(assochildren[26])),removeapro(str(assochildren[27])),removeapro(str(assochildren[28])),
                        removeapro(str(assochildren[29])),removeapro(str(assochildren[30])),removeapro(str(assochildren[31])),
                        removeapro(str(assochildren[32])),removeapro(str(assochildren[33])),removeapro(str(assochildren[34])),
                        removeapro(str(assochildren[35])),removeapro(str(assochildren[36])),removeapro(str(assochildren[37])),
                        removeapro(str(assochildren[38])),removeapro(str(assochildren[39])),removeapro(str(assochildren[40])),
                        removeapro(str(assochildren[41])),removeapro(str(assochildren[42])),removeapro(str(assochildren[43])),
                        removeapro(str(assochildren[44]))                        
                        
                        ))
                        con.commit()

            # spoouse
            relNames='spouse'
            if(spouse=='None'):
                spouse=''
            else:
                find_it = "and"
                repl_it = ","
                spouse = spouse.replace(find_it, repl_it)
                spouse = spouse.replace("&", "")
                spouse=spouse.split(",")
                if len(spouse) > 1:
                    for i in spouse:
                        p=i
                        i=i.split(' ')
                        if len(i) > 1:
                            scib_firstName=i[0]
                            scib_middleName=i[1]
                            if scib_firstName in array1:
                                scib_title=i[0]
                                scib_firstName=i[1]
                            else:
                                scib_firstName=i[0]
                            if scib_firstName=='':
                                scib_firstName=i[1]
                            else:
                                scib_firstName=scib_firstName
                        else:
                            scib_firstName=i
                        scib_middleName=str(scib_middleName).split(' ')
                        if len(scib_middleName)>1:
                            scib_middleName=scib_middleName[0]
                            scib_lastName=scib_middleName[1]
                        else:
                            scib_lastName=scib_middleName[0]
                        
                        scib_middleName=str(scib_middleName).removeprefix("['")
                        scib_middleName=str(scib_middleName).removesuffix("']")
                        scib_middleName=str(scib_middleName).removeprefix('["')
                        spouse=str(scib_middleName).removesuffix('"]')  

                        scib_lastName=str(scib_lastName).removeprefix("['")
                        scib_lastName=str(scib_lastName).removesuffix("']")
                        scib_lastName=str(scib_lastName).removeprefix('["')
                        scib_lastName=str(scib_lastName).removesuffix('"]')

                        scib_firstName=str(scib_firstName).removeprefix("['")
                        scib_firstName=str(scib_firstName).removesuffix("']")
                        scib_firstName=str(scib_firstName).removeprefix('["')
                        scib_firstName=str(scib_firstName).removesuffix('"]')

                        p=str(p).removeprefix("['")
                        p=str(p).removesuffix("']")
                        p=str(p).removeprefix('["')
                        p=str(p).removesuffix('"]')

                        FirstNames =nam(p)
                     #print(ic4proFirstNames)

                        Surnames =last_name(p)
                     #print(ic4proSurnames)
                        if scib_firstName in titlearray:
                            with open('firstswap.txt', 'a') as f:
                                f.write(str(radproSurname)+"\n")
                        watch=watchId(p)
                        radproWatchlistId =  radproCategory+"-"+radproFirstName +'-'+ watchID
                        radproAssociateId =radproCategory+"-"+  scib_firstName+"-"+radproFirstName +'-'+ watchID

                        if str(scib_middleName)==str(scib_lastName):
                            scib_middleName=''
                        with open('souse.txt', 'a') as m:
                            m.write("Middle Name:"+str(scib_middleName)+"\n")
                            m.write("Firstname Name:"+str(FirstNames)+"\n")
                            m.write("Firstname f:"+str(scib_firstName)+"\n")
                            m.write("Surname Name:"+str(Surnames)+"\n")

                        assospouse=(
                        replacespace(removedoublehyphen (radproAssociateId.upper())),
                        replacespace(removedoublehyphen(radproWatchlistId.upper())),	
                        removing_none(scib_title),
                        removing_none(without_non_alphabets (str(scib_firstName))),
                        removing_none(FirstNames),
                        removing_none(Surnames),
                        removing_none(radproMaidenName),	
                        removing_none(radproOtherNames),
                        without_non_alphabet(str(p)),	
                        removing_none(relNames),	
                        removing_none(radproType),	
                        removing_none(radproAction),	
                        replaceslash(removing_none(radproActionDate)),	
                        removing_none(radproGender),	
                        removing_none(radproStatus),	
                        replaceslash(removing_none(radproDOB)),	
                        removing_none(radproPOB),	
                        removing_none(radproAddress),	
                        removing_none(radproCity),	
                        removing_none(radproCounty),	
                        removing_none(radproPostalCode),	
                        removing_none(radproState),	
                        removing_none(radproCountry),	
                        removing_none(radproCountryCode),	
                        removing_none(radproContinent),	
                        removing_none(radproCategory),	
                        removing_none(radproSubCategory),	
                        removing_none(radproWatchType),	
                        removing_none(radproBranch),	
                        removing_none(radproCluster),	
                        removing_none(radproZone),
                        removing_none(radproRegion),	
                        removing_none(radproImageLink),	
                        removing_none(radproUrl),	
                        removing_none(radproTier),	
                        removing_none(radproRisk),	
                        removing_none(radproTierRank),	
                        removing_none(radproDescription),	
                        removing_none(radproLanguage),	
                        str(radproRecordDate),	
                        removing_none(radproRecordTime),	
                        removing_none(radproOperator),	
                        removing_none(radproOperation),	
                        removing_none(radproWorkstation),	
                        removing_none(radproRecordCounter)
                        )
                        cursor.execute('''
                        SELECT * FROM U_IC4INDEP.PEP_ASSOCIATES WHERE RADPROASSOCIATEID='{0}'
                    '''.format(str(replacespace(removedoublehyphen(radproAssociateId.upper())))))
                        row = cursor.fetchone()
                        if not row:
                            cursor.execute('''
                        INSERT INTO "U_IC4INDEP"."PEP_ASSOCIATES" 
                        (
                                "RADPROASSOCIATEID" , 
                                "RADPROWATCHLISTID" , 
                                "RADPROTITLE", 
                                "RADPROFIRSTNAME" , 
                                "RADPROMIDDLENAME" , 
                                "RADPROSURNAME" , 
                                "RADPROMAIDENNAME" , 
                                "RADPROOTHERNAMES" , 
                                "RADPROFULLNAME" , 
                                "RADPRORELNAME" , 
                                "RADPROTYPE" , 
                                "RADPROACTION" , 
                                "RADPROACTIONDATE" , 
                                "RADPROGENDER" , 
                                "RADPROSTATUS" , 
                                "RADPRODOB" , 
                                "RADPROPOB", 
                                "RADPROADDRESS", 
                                "RADPROCITY" , 
                                "RADPROCOUNTY" , 
                                "RADPROPOSTALCODE" , 
                                "RADPROSTATE" , 
                                "RADPROCOUNTRY" , 
                                "RADPROCOUNTRYCODE" , 
                                "RADPROCONTINENT" , 
                                "RADPROCATEGORY" , 
                                "RADPROSUBCATEGORY" , 
                                "RADPROWATCHTYPE" , 
                                "RADPROBRANCH" , 
                                "RADPROCLUSTER" , 
                                "RADPROZONE" , 
                                "RADPROREGION" , 
                                "RADPROIMAGELINK" , 
                                "RADPROURL" , 
                                "RADPROTIER" , 
                                "RADPRORISK" , 
                                "RADPROTIERRANK" , 
                                "RADPRODESCRIPTION" , 
                                "RADPROLANGUAGE" , 
                                "RADPRORECORDDATE" , 
                                "RADPRORECORDTIME" , 
                                "RADPROOPERATOR" , 
                                "RADPROOPERATION" , 
                                "RADPROWORKSTATION" , 
                                "RADPRORECORDCOUNTER"

                        )
                        VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}',
                                '{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
                                '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}','{30}','{31}','{32}',
                                '{33}','{34}','{35}','{36}','{37}','{38}','{39}','{40}','{41}','{42}','{43}',
                                '{44}'
                        )
                        '''.format(
                        removeapro(str(assospouse[0])), removeapro(str(assospouse[1])), removeapro(str(assospouse[2])),
                        removeapro(str(assospouse[3])), removeapro(str(assospouse[4])), removeapro(str(assospouse[5])), 
                        removeapro(str(assospouse[6])), removeapro(str(assospouse[7])), removeapro(str(assospouse[8])), 
                        removeapro(str(assospouse[9])), removeapro(str(assospouse[10])), removeapro(str(assospouse[11])), 
                        removeapro(str(assospouse[12])), removeapro(str(assospouse[13])), removeapro(str(assospouse[14])), 
                        replaceslash(removeapro(str(assospouse[15]))), removeapro(str(assospouse[16])),
                        removeapro(str(assospouse[17])),removeapro(str(assospouse[18])),removeapro(str(assospouse[19])),
                        removeapro(str(assospouse[20])),removeapro(str(assospouse[21])),removeapro(str(assospouse[22])),removeapro(str(assospouse[23])),removeapro(str(assospouse[24])),
                        removeapro(str(assospouse[25])),
                        removeapro(str(assospouse[26])),removeapro(str(assospouse[27])),removeapro(str(assospouse[28])),
                        removeapro(str(assospouse[29])),removeapro(str(assospouse[30])),removeapro(str(assospouse[31])),
                        removeapro(str(assospouse[32])),removeapro(str(assospouse[33])),removeapro(str(assospouse[34])),
                        removeapro(str(assospouse[35])),removeapro(str(assospouse[36])),removeapro(str(assospouse[37])),
                        removeapro(str(assospouse[38])),removeapro(str(assospouse[39])),removeapro(str(assospouse[40])),
                        removeapro(str(assospouse[41])),removeapro(str(assospouse[42])),removeapro(str(assospouse[43])),
                        removeapro(str(assospouse[44]))                        
                        
                        ))
                            con.commit()
                else:
                    spouse=str(spouse).removeprefix("['")
                    spouse=spouse.removesuffix("']")
                    spouse=str(spouse).removeprefix('["')
                    spouse=spouse.removesuffix('"]')  
                    scibb=spouse
                    spouse=spouse.split(' ')

                    if len(spouse) > 1:
                        scib_firstName=spouse[0]
                        # print('cib_title:',cib_title)
                        scib_middleName=spouse[1]
                        if scib_firstName in array1:
                           scib_title=spouse[0]
                           scib_firstName=spouse[1]
                        else:
                            scib_firstName=spouse[0]
                        if scib_firstName=='':
                            scib_firstName=spouse[1]
                        else:
                            scib_firstName=scib_firstName
                    else:
                        scib_firstName=spouse
                    scib_middleName=str(scib_middleName).split(' ')
                    if len(scib_middleName)>1:
                        scib_middleName=scib_middleName[0]
                        scib_lastName=scib_middleName[1]
                    else:
                        scib_lastName=scib_middleName[0]
                    
                    scib_middleName=str(scib_middleName).removeprefix("['")
                    scib_middleName=str(scib_middleName).removesuffix("']")
                    scib_middleName=str(scib_middleName).removeprefix('["')
                    spouse=str(scib_middleName).removesuffix('"]')  

                    scib_lastName=str(scib_lastName).removeprefix("['")
                    scib_lastName=str(scib_lastName).removesuffix("']")
                    scib_lastName=str(scib_lastName).removeprefix('["')
                    scib_lastName=str(scib_lastName).removesuffix('"]')

                    scib_lastName=str(scib_lastName).removeprefix("['")
                    scib_lastName=str(scib_lastName).removesuffix("']")
                    scib_lastName=str(scib_lastName).removeprefix('["')
                    scib_lastName=str(scib_lastName).removesuffix('"]')

                    scibb=str(scibb).removeprefix("['")
                    scibb=str(scibb).removesuffix("']")
                    scibb=str(scibb).removeprefix('["')
                    scibb=str(scibb).removesuffix('"]')

                    ic4pro =nam(scibb)
                    #print(ic4proFirstNames)

                  
                    ic4proS =last_name(scibb)
                  #print(ic4proSurnames)
                    # radproAssociateId =radproCategory+"-"+ without_non_alphabets (str(cib_firstName))+"-"+radproFirstName +'-'+ watchID

                    watchID=watchId(scibb)
                    radproAssociateId = radproCategory+"-"+ without_non_alphabets(str(ic4proS))+"-"+radproFirstName+'-'+watchID
                    radproWatchlistId =radproCategory+"-"+radproFirstName +'-'+ watchID

                    if str(scib_middleName)==str(scib_lastName):
                        scib_middleName='' 
                    ID = removedoublehyphen(radproWatchlistId.upper())
                    
                    spounse2=(
                        replacespace(removedoublehyphen (radproAssociateId.upper())),
                        replacespace(removedoublehyphen(radproWatchlistId.upper())),	
                        removing_none(scib_title),
                        removing_none(without_non_alphabets (str(scib_firstName))),
                        removing_none(FirstNames),
                        removing_none(ic4proS),
                        removing_none(radproMaidenName),	
                        removing_none(radproOtherNames),
                        without_non_alphabet(str(scibb)),	
                        removing_none(relNames),	
                        removing_none(radproType),	
                        removing_none(radproAction),	
                        replaceslash(removing_none(radproActionDate)),	
                        removing_none(radproGender),	
                        removing_none(radproStatus),	
                        replaceslash(removing_none(radproDOB)),	
                        removing_none(radproPOB),	
                        removing_none(radproAddress),	
                        removing_none(radproCity),	
                        removing_none(radproCounty),	
                        removing_none(radproPostalCode),	
                        removing_none(radproState),	
                        removing_none(radproCountry),	
                        removing_none(radproCountryCode),	
                        removing_none(radproContinent),	
                        removing_none(radproCategory),	
                        removing_none(radproSubCategory),	
                        removing_none(radproWatchType),	
                        removing_none(radproBranch),	
                        removing_none(radproCluster),	
                        removing_none(radproZone),
                        removing_none(radproRegion),	
                        removing_none(radproImageLink),	
                        removing_none(radproUrl),	
                        removing_none(radproTier),	
                        removing_none(radproRisk),	
                        removing_none(radproTierRank),	
                        removing_none(radproDescription),	
                        removing_none(radproLanguage),	
                        str(radproRecordDate),	
                        removing_none(radproRecordTime),	
                        removing_none(radproOperator),	
                        removing_none(radproOperation),	
                        removing_none(radproWorkstation),	
                        removing_none(radproRecordCounter)
                    )
                    cursor.execute('''
                        SELECT * FROM U_IC4INDEP.PEP_ASSOCIATES WHERE RADPROASSOCIATEID='{0}'
                    '''.format(str(removeapro (replacespace(removedoublehyphen(radproAssociateId.upper()))))))
                    row = cursor.fetchone()
                    if not row:
                        cursor.execute('''
                        INSERT INTO "U_IC4INDEP"."PEP_ASSOCIATES" 
                        (
                                "RADPROASSOCIATEID" , 
                                "RADPROWATCHLISTID" , 
                                "RADPROTITLE", 
                                "RADPROFIRSTNAME" , 
                                "RADPROMIDDLENAME" , 
                                "RADPROSURNAME" , 
                                "RADPROMAIDENNAME" , 
                                "RADPROOTHERNAMES" , 
                                "RADPROFULLNAME" , 
                                "RADPRORELNAME" , 
                                "RADPROTYPE" , 
                                "RADPROACTION" , 
                                "RADPROACTIONDATE" , 
                                "RADPROGENDER" , 
                                "RADPROSTATUS" , 
                                "RADPRODOB" , 
                                "RADPROPOB", 
                                "RADPROADDRESS", 
                                "RADPROCITY" , 
                                "RADPROCOUNTY" , 
                                "RADPROPOSTALCODE" , 
                                "RADPROSTATE" , 
                                "RADPROCOUNTRY" , 
                                "RADPROCOUNTRYCODE" , 
                                "RADPROCONTINENT" , 
                                "RADPROCATEGORY" , 
                                "RADPROSUBCATEGORY" , 
                                "RADPROWATCHTYPE" , 
                                "RADPROBRANCH" , 
                                "RADPROCLUSTER" , 
                                "RADPROZONE" , 
                                "RADPROREGION" , 
                                "RADPROIMAGELINK" , 
                                "RADPROURL" , 
                                "RADPROTIER" , 
                                "RADPRORISK" , 
                                "RADPROTIERRANK" , 
                                "RADPRODESCRIPTION" , 
                                "RADPROLANGUAGE" , 
                                "RADPRORECORDDATE" , 
                                "RADPRORECORDTIME" , 
                                "RADPROOPERATOR" , 
                                "RADPROOPERATION" , 
                                "RADPROWORKSTATION" , 
                                "RADPRORECORDCOUNTER"

                        )
                        VALUES ('{0}','{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}',
                                '{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
                                '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}','{30}','{31}','{32}',
                                '{33}','{34}','{35}','{36}','{37}','{38}','{39}','{40}','{41}','{42}','{43}',
                                '{44}'
                        )
                        '''.format(
                        removeapro(str(spounse2[0])), removeapro(str(spounse2[1])), removeapro(str(spounse2[2])),
                        removeapro(str(spounse2[3])), removeapro(str(spounse2[4])), removeapro(str(spounse2[5])), 
                        removeapro(str(spounse2[6])), removeapro(str(spounse2[7])), removeapro(str(spounse2[8])), 
                        removeapro(str(spounse2[9])), removeapro(str(spounse2[10])), removeapro(str(spounse2[11])), 
                        removeapro(str(spounse2[12])), removeapro(str(spounse2[13])), removeapro(str(spounse2[14])), 
                        replaceslash(removeapro(str(spounse2[15]))), removeapro(str(spounse2[16])),
                        removeapro(str(spounse2[17])),removeapro(str(spounse2[18])),removeapro(str(spounse2[19])),
                        removeapro(str(spounse2[20])),removeapro(str(spounse2[21])),removeapro(str(spounse2[22])),removeapro(str(spounse2[23])),removeapro(str(spounse2[24])),
                        removeapro(str(spounse2[25])),
                        removeapro(str(spounse2[26])),removeapro(str(spounse2[27])),removeapro(str(spounse2[28])),
                        removeapro(str(spounse2[29])),removeapro(str(spounse2[30])),removeapro(str(spounse2[31])),
                        removeapro(str(spounse2[32])),removeapro(str(spounse2[33])),removeapro(str(spounse2[34])),
                        removeapro(str(spounse2[35])),removeapro(str(spounse2[36])),removeapro(str(spounse2[37])),
                        removeapro(str(spounse2[38])),removeapro(str(spounse2[39])),removeapro(str(spounse2[40])),
                        removeapro(str(spounse2[41])),removeapro(str(spounse2[42])),removeapro(str(spounse2[43])),
                        removeapro(str(spounse2[44]))                        
                        
                        ))
                        con.commit()


                  
                        



            testingdic = (
                    replacespace(radproWatchlistId.upper()),
                    removing_none(radproTitle),
                    removing_none(radproFirstName),
                    removing_none(radproMiddleName),
                    removing_none(radproSurname),
                    removing_none(remnone(radproFullName)),	
                    removing_none(radproAction),	
                    removing_none(radproActionDate),	
                    removing_none(radproGender),	
                    removing_none(radproStatus),	
                    removing_none(radproDOB),
                    removing_none(radproPOB),

                    removing_none(radproAltDOB),
                    removing_none(radproAddress),
                    removing_none(radproCity),
                    removing_none(radproCounty),	
                    removing_none(radproPostalCode),	
                    removing_none(radproState),
                    removing_none(radproCountry),
                    removing_none(radproCountryCode),	
                    removing_none(radproContinent),
                    removing_none(radproCategory),
                    removing_none(radproSubCategory),	
                    removing_none(radproWatchType),
                    removing_none(radproBlacklistedId),
                    removing_none(radproReportedBy),
                    removing_none(radproPortName),
                    removing_none(radproPortType),
                    removing_none(radproWhiteFlag),
                    removing_none(radproFlagDate),
                    removing_none(radproFlagTime),
                    removing_none(radproReason),
                    removing_none(radproFlagBy),
                    removing_none(radproBranch),
                    removing_none(radproCluster),
                    removing_none(radproZone),
                    removing_none(radproRegion),	
                    removing_none(radproRemarks),	
                    removing_none(radproImageLink),
                    removing_none(radproUrl),
                    removing_none(radproTags),
                    removing_none(radproTier),
                    removing_none(radproTierRank),	
                    removing_none(radproProfession),	
                    removing_none(radproDescription),	
                    removing_none(radproLanguage),	
                    removing_none(str(radproRecordDate)),	
                    removing_none(radproRecordTime),	
                    removing_none(radproOperator),	
                    removing_none(radproOperation),
                    removing_none(radproWorkstation),
                    removing_none(radproRecordCounter)
                    )
            testing = str(testingdic).removesuffix(")")
            # cursor.execute('''
            #         SELECT * FROM U_IC4INDEP.PEP_WATCHLIST WHERE RADPROWATCHLISTID='{0}'
            #         '''.format(str(removeapro(removedoublehyphen(radproWatchlistId.upper()))) ))
            # row = cursor.fetchone()
            # if not row:
            #     cursor.execute('''
            #                 INSERT INTO U_IC4INDEP.PEP_WATCHLIST (
            #                "RADPROWATCHLISTID", 
            #                 "RADPROTITLE", 
            #                 "RADPROFIRSTNAME", 
            #                 "RADPROMIDDLENAME", 
            #                 "RADPROSURNAME", 
            #                 "RADPROFULLNAME", 
            #                 "RADPROACTION",
            #                 "RADPROACTIONDATE", 
            #                 "RADPROGENDER", 
            #                 "RADPROSTATUS", 
            #                 "RADPRODOB", 
            #                 "RADPROALTDOB", 
            #                 "RADPROPOB", 
            #                 "RADPROADDRESS", 
            #                 "RADPROCITY", 
            #                 "RADPROCOUNTY", 
            #                 "RADPROPOSTALCODE", 
            #                 "RADPROSTATE", 
            #                 "RADPROCOUNTRY", 
            #                 "RADPROCOUNTRYCODE", 
            #                 "RADPROCONTINENT", 
            #                 "RADPROCATEGORY", 
            #                 "RADPROSUBCATEGORY", 
            #                 "RADPROWATCHTYPE", 
            #                 "RADPROBLACKLISTEDID", 
            #                 "RADPROREPORTEDBY", 
            #                 "RADPROPORTNAME", 
            #                 "RADPROPORTTYPE", 
            #                 "RADPROWHITEFLAG", 
            #                 "RADPROFLAGDATE", 
            #                 "RADPROFLAGTIME", 
            #                 "RADPROREASON", 
            #                 "RADPROFLAGBY", 
            #                 "RADPROBRANCH", 
            #                 "RADPROCLUSTER", 
            #                 "RADPROZONE", 
            #                 "RADPROREGION", 
            #                 "RADPROREMARKS", 
            #                 "RADPROIMAGELINK", 
            #                 "RADPROURL", 
            #                 "RADPROTAGS", 
            #                 "RADPROTIER", 
            #                 "RADPROTIERRANK", 
            #                 "RADPROPROFESSION", 
            #                 "RADPRODESCRIPTION", 
            #                 "RADPROLANGUAGE", 
            #                 "RADPRORECORDDATE", 
            #                 "RADPRORECORDTIME", 
            #                 "RADPROOPERATOR", 
            #                 "RADPROOPERATION", 
            #                 "RADPROWORKSTATION", 
            #                 "RADPRORECORDCOUNTER"
            #             )
            #             VALUES ('{0}', '{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}',
            #                 '{10}','{11}','{12}','{13}','{14}','{15}','{16}','{17}','{18}','{19}','{20}','{21}',
            #                 '{22}','{23}','{24}','{25}','{26}','{27}','{28}','{29}',
            #                 '{30}','{31}','{32}','{33}','{34}','{35}','{36}','{37}',
            #                 '{38}','{39}','{40}','{41}','{42}','{43}','{44}','{45}','{46}',
            #                 '{47}', '{48}', '{49}', '{50}', '{51}')
            #             '''.format(removeapro(str(testingdic[0])), removeapro(str(testingdic[1])), removeapro(str(testingdic[2])), removeapro(str(testingdic[3])), 
            #             removeapro(str(testingdic[4])),
            #             removeapro(str(testingdic[5])), removeapro(str(testingdic[6])), removeapro(str(testingdic[7])), 
            #             removeapro(str(testingdic[8])), removeapro(str(testingdic[9])), replaceslash( removeapro(str(testingdic[10]))),
            #             removeapro(str(testingdic[11])), replaceslash(removeapro(str(testingdic[12]))), removeapro(str(testingdic[13])), removeapro(str(testingdic[14])), 
            #             removeapro(str(testingdic[15])), removeapro(str(testingdic[16])),
            #             removeapro(str(testingdic[17])), removeapro(str(testingdic[18])), removeapro(str(testingdic[19])), removeapro(str(testingdic[20])), removeapro(str(testingdic[21])), removeapro(str(testingdic[22])),
            #             removeapro(str(testingdic[23])), removeapro(str(testingdic[24])), removeapro(str(testingdic[25])), removeapro(str(testingdic[26])), removeapro(str(testingdic[27])), removeapro(str(testingdic[28])),
            #             replaceslash(removeapro(str(testingdic[29]))), removeapro(str(testingdic[30])), removeapro(str(testingdic[31])), removeapro(str(testingdic[32])), removeapro(str(testingdic[33])), removeapro(str(testingdic[34])),
            #             removeapro(str(testingdic[35])), removeapro(str(testingdic[36])), removeapro(str(testingdic[37])), removeapro(str(testingdic[38])), removeapro(str(testingdic[39])), removeapro(str(testingdic[40])),
            #             removeapro(str(testingdic[41])), removeapro(str(testingdic[42])), removeapro(str(testingdic[43])), removeapro(str(testingdic[44])), removeapro(str(testingdic[45])), removeapro(str(testingdic[46])),
            #             removeapro(str(testingdic[47])), removeapro(str(testingdic[48])), removeapro(str(testingdic[49])), removeapro(str(testingdic[50])), testingdic[51]
            #             ))

            #     con.commit()




            # #    with open('top.txt', 'a') as f:
            # #     f.write(str(ic4proFullName))
            # #     f.write('\n')
            # #    with open('row1.txt', 'a') as f:
            # #         f.write(str(cell_obj))
            # #         f.write('\n')
